import random
import string
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

def generate_random_string(min_length, max_length):
    length = random.randint(min_length, max_length)
    letters_and_digits = string.ascii_letters
    return ''.join(random.choice(letters_and_digits) for _ in range(length))

def generate_random_date():
    start_date = datetime(1950, 1, 1)
    end_date = datetime(2000, 12, 31)
    random_date = start_date + timedelta(days=random.randint(0, (end_date - start_date).days))
    return random_date

def generate_data():
    account_name = generate_random_string(6, 10)
    password = generate_random_string(8, 15) + str(random.randint(0, 9))

    # Generating a name with length greater than 6
    first_name = generate_random_string(7, 16)

    birth_date = generate_random_date()
    day, month, year = str(birth_date.day).zfill(2), str(birth_date.month).zfill(2), str(birth_date.year)

    name_suffix = str(random.randint(1000, 9999))

    return account_name, password, day, month, year, first_name, name_suffix

def save_to_excel(data):
    wb = Workbook()
    ws = wb.active

    # Adding a text format style
    text_format = NamedStyle(name='text_format', number_format='@')
    wb.add_named_style(text_format)

    # Adding title row
    ws.append(["Account Name", "Password", "Day", "Month", "Year", "First Name", "Name Suffix"])

    # Adding data
    for row in data:
        ws.append(row)

    # Applying text format to all cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = 'text_format'

    # Save the Excel file
    wb.save("generated_data_updated_text_format.xlsx")

if __name__ == "__main__":
    # Generate 300 pieces of data
    generated_data = [generate_data() for _ in range(300)]

    # Save data to Excel
    save_to_excel(generated_data)

    print("Data generation and Excel file creation complete.")
